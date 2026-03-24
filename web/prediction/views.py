import csv
import io
import json
import os
import sys
from datetime import datetime

import pandas as pd
from django.conf import settings
from django.http import HttpResponse
from django.shortcuts import redirect, render

from .forms import PredictionForm

# ─────────────────────────────────────────────────────────────
# 表單欄位名稱 → 模型 DataFrame 欄位名稱的映射
# ─────────────────────────────────────────────────────────────
FORM_TO_MODEL_MAPPING = {
    "education": "education",
    "month_salary": "month salary",
    "job_tenure": "job tenure",
    "residence_status": "residence status",
    "main_business": "main business",
    "product": "product",
    "loan_term": "loan term",
    "paid_installments": "paid installments",
    "debt_to_income_ratio": "debt_to_income_ratio",
    "payment_to_income_ratio": "payment_to_income_ratio",
    "post_code_permanent": "post code of permanent address",
    "post_code_residential": "post code of residential address",
    "overdue_before_first": "number of overdue before the first month",
    "overdue_first_half": "number of overdue in the first half of the first month",
    "overdue_first_second_half": "number of overdue in the second half of the first month",
    "overdue_month_2": "number of overdue in the second month",
    "overdue_month_3": "number of overdue in the third month",
    "overdue_month_4": "number of overdue in the fourth month",
    "overdue_month_5": "number of overdue in the fifth month",
    "overdue_month_6": "number of overdue in the sixth month",
}

# ─────────────────────────────────────────────────────────────
# 載入 DPMPredictor（僅在首次呼叫時初始化）
# ─────────────────────────────────────────────────────────────
_predictor = None


def _get_predictor():
    """延遲載入 DPMPredictor 單例。"""
    global _predictor
    if _predictor is not None:
        return _predictor

    # 將專案根目錄加入 sys.path
    project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
    prediction_dir = os.path.join(project_root, "Prediction")
    if prediction_dir not in sys.path:
        sys.path.insert(0, prediction_dir)

    # 切換工作目錄以便相對路徑正確解析
    original_cwd = os.getcwd()
    os.chdir(prediction_dir)
    try:
        from predict import DPMPredictor
        _predictor = DPMPredictor()
    finally:
        os.chdir(original_cwd)

    return _predictor


def _form_to_dataframe(cleaned_data: dict) -> pd.DataFrame:
    """將 Django 表單清洗後的資料轉為模型所需的單行 DataFrame。"""
    row = {}
    for form_field, model_col in FORM_TO_MODEL_MAPPING.items():
        row[model_col] = cleaned_data[form_field]
    return pd.DataFrame([row])


# ─────────────────────────────────────────────────────────────
# 模型預測
# ─────────────────────────────────────────────────────────────
NTD_TO_USD = 0.031  # approximate NTD→USD conversion rate

# ─────────────────────────────────────────────────────────────
# English → Chinese translation for uploaded batch data
# ─────────────────────────────────────────────────────────────
EN_TO_ZH = {
    # education
    "Grad+": "研究所以上", "Graduate": "研究所以上", "Master": "碩士",
    "College/Univ": "專科/大學", "University": "大學", "College": "專科",
    "HS/Voc": "高中/職", "High School": "高中", "Elementary": "小學",
    "Other": "其他", "Others": "其他",
    # residence status
    "Owned": "自有", "Rented": "租屋", "Spouse": "配偶",
    "Family": "親屬", "Dormitory": "宿舍",
    # main business
    "Manufacturing": "製造業", "Services": "服務業", "Service": "服務業",
    "Commerce": "商業", "Retail": "商業",
    "Tech": "科技業", "Finance": "金融業", "Insurance": "保險業",
    "Securities": "證券及期貨業", "Gov/Education": "公教人員", "Military": "軍人",
    "F&B": "餐飲業", "Transport": "運輸業", "Construction": "營造業",
    "Real Estate": "不動產業", "Warehousing": "倉儲業", "Telecom": "通信業",
    "Utilities": "水電燃氣業", "Agriculture": "農牧林漁", "Fishery": "漁業",
    "Mining": "礦業及土石採取業", "Professional": "專業人士", "Freelance": "自由業",
    "E-commerce": "網拍業", "Student": "學生", "Homemaker": "家管",
    "Social/Personal": "社會團體即個人服務",
    # product
    "Beauty": "瘦身美容", "3C/Appliance": "3C家電", "Personal": "個人用品",
    "Personal Items": "個人用品",
}

# Chinese → English translation for model output values
ZH_TO_EN_GRADE = {
    "A (優良)": "A (Excellent)", "B (中等)": "B (Moderate)",
    "C (警戒)": "C (Warning)", "D (不良)": "D (Poor)", "E (危險)": "E (Critical)",
}
ZH_TO_EN_ACTION = {
    "正常監控": "Normal Monitoring", "關注名單": "Watch List",
    "早期預警": "Early Warning", "加強監控": "Enhanced Monitoring",
    "惡化警示": "Deterioration Alert", "立即介入": "Immediate Intervention",
    "強制催收": "Forced Collection", "加強監控+早期預警": "Enhanced Monitoring + Early Warning",
    "加強監控/催收準備": "Enhanced Monitoring / Collection Prep",
    "立即催收": "Immediate Collection",
    "立即介入/催收準備": "Immediate Intervention / Collection Prep",
}
# Columns in uploaded data that may contain English categorical values
EN_CATEGORICAL_COLS = ["education", "residence status", "main business", "product"]
# Reverse mapping: Chinese → English (for translating form values in output)
ZH_TO_EN = {v: k for k, v in EN_TO_ZH.items()}


def predict_model(cleaned_data: dict, lang: str = "zh") -> dict:
    """
    接收表單清洗後的資料，呼叫真實 DPMPredictor 回傳預測結果。
    lang: "zh" or "en" — controls primary display language and currency.
    Always returns both _zh and _en variants for client-side toggle support.
    """
    predictor = _get_predictor()

    original_salary = cleaned_data.get("month_salary", 0)
    # Determine salary in both currencies
    if lang == "en":
        cleaned_data = dict(cleaned_data)  # avoid mutating original
        cleaned_data["month_salary"] = original_salary / NTD_TO_USD
        salary_usd = original_salary
        salary_ntd = original_salary / NTD_TO_USD
    else:
        salary_ntd = original_salary
        salary_usd = original_salary * NTD_TO_USD

    df = _form_to_dataframe(cleaned_data)

    # 呼叫模型
    result_df = predictor.predict_with_details(df, simplified_output=False)
    row = result_df.iloc[0]

    prob = float(row["default_probability"])       # 已是百分比 0-100
    risk_score = int(row["risk_score"])             # 0-100
    risk_grade_raw = str(row["risk_grade"])         # always zh from model, e.g. "A (優良)"
    risk_alert = str(row["risk_alert"])             # always EN from model, e.g. "LOW RISK"
    recommendation_raw = str(row.get("risk_action_optimal", row.get("risk_action", "")))  # always zh

    grade_letter = risk_grade_raw[0] if risk_grade_raw else "C"
    color_map = {"A": "emerald", "B": "green", "C": "amber", "D": "orange", "E": "red"}
    risk_color = color_map.get(grade_letter, "gray")

    # Both language variants
    risk_grade_zh = risk_grade_raw
    risk_grade_en = ZH_TO_EN_GRADE.get(risk_grade_raw, risk_grade_raw)
    risk_label_zh = {"A": "極低風險", "B": "中等風險", "C": "高風險", "D": "極高風險", "E": "危急風險"}.get(grade_letter, "未知")
    risk_label_en = {"A": "Very Low Risk", "B": "Moderate Risk", "C": "High Risk", "D": "Very High Risk", "E": "Critical Risk"}.get(grade_letter, "Unknown")
    recommendation_zh = recommendation_raw
    recommendation_en = ZH_TO_EN_ACTION.get(recommendation_raw, recommendation_raw)

    _t = lambda v: ZH_TO_EN.get(str(v), v)  # translate ZH categorical → EN

    overdue_total = str(sum(
        cleaned_data.get(f, 0) for f in [
            "overdue_before_first", "overdue_first_half",
            "overdue_first_second_half", "overdue_month_2",
            "overdue_month_3", "overdue_month_4",
            "overdue_month_5", "overdue_month_6",
        ]
    ))

    # Bilingual rows for client-side toggle: [{key_zh, key_en, val_zh, val_en}, ...]
    input_summary_rows = [
        {"key_zh": "教育程度",    "key_en": "Education",              "val_zh": cleaned_data.get("education", ""),          "val_en": _t(cleaned_data.get("education", ""))},
        {"key_zh": "月薪",        "key_en": "Monthly Salary",         "val_zh": f'{salary_ntd:,.0f} 元',                    "val_en": f'${salary_usd:,.0f} USD'},
        {"key_zh": "工作年資",    "key_en": "Job Tenure",             "val_zh": f'{cleaned_data.get("job_tenure", 0)} 年',   "val_en": f'{cleaned_data.get("job_tenure", 0)} years'},
        {"key_zh": "居住狀態",    "key_en": "Residence Status",       "val_zh": cleaned_data.get("residence_status", ""),   "val_en": _t(cleaned_data.get("residence_status", ""))},
        {"key_zh": "行業別",      "key_en": "Industry",               "val_zh": cleaned_data.get("main_business", ""),      "val_en": _t(cleaned_data.get("main_business", ""))},
        {"key_zh": "借款目的",    "key_en": "Loan Purpose",           "val_zh": cleaned_data.get("product", ""),            "val_en": _t(cleaned_data.get("product", ""))},
        {"key_zh": "貸款期數",    "key_en": "Loan Term",              "val_zh": f'{cleaned_data.get("loan_term", 0)} 期',    "val_en": f'{cleaned_data.get("loan_term", 0)} months'},
        {"key_zh": "已繳期數",    "key_en": "Paid Installments",      "val_zh": f'{cleaned_data.get("paid_installments", 0)} 期', "val_en": f'{cleaned_data.get("paid_installments", 0)} months'},
        {"key_zh": "負債收入比",  "key_en": "Debt-to-Income Ratio",   "val_zh": f'{cleaned_data.get("debt_to_income_ratio", 0)}',   "val_en": f'{cleaned_data.get("debt_to_income_ratio", 0)}'},
        {"key_zh": "還款收入比",  "key_en": "Payment-to-Income Ratio","val_zh": f'{cleaned_data.get("payment_to_income_ratio", 0)}', "val_en": f'{cleaned_data.get("payment_to_income_ratio", 0)}'},
        {"key_zh": "戶籍郵遞區號","key_en": "Perm. Postal Code",      "val_zh": str(cleaned_data.get("post_code_permanent", "")),   "val_en": str(cleaned_data.get("post_code_permanent", ""))},
        {"key_zh": "居住郵遞區號","key_en": "Res. Postal Code",       "val_zh": str(cleaned_data.get("post_code_residential", "")), "val_en": str(cleaned_data.get("post_code_residential", ""))},
        {"key_zh": "逾期總次數",  "key_en": "Total Overdue Count",    "val_zh": overdue_total,                              "val_en": overdue_total},
    ]

    # Primary display based on lang
    risk_grade = risk_grade_en if lang == "en" else risk_grade_zh
    risk_label = risk_label_en if lang == "en" else risk_label_zh
    recommendation = recommendation_en if lang == "en" else recommendation_zh
    key_lang = "key_en" if lang == "en" else "key_zh"
    val_lang = "val_en" if lang == "en" else "val_zh"
    input_summary = {r[key_lang]: r[val_lang] for r in input_summary_rows}

    return {
        "default_probability": round(prob, 2),
        "risk_score": risk_score,
        "risk_grade": risk_grade,
        "risk_label": risk_label,
        "risk_color": risk_color,
        "risk_alert": risk_alert,
        "recommendation": recommendation,
        "input_summary": input_summary,        # for CSV/Excel download (lang-specific)
        # Both language variants for client-side toggle
        "risk_grade_zh": risk_grade_zh,
        "risk_grade_en": risk_grade_en,
        "risk_label_zh": risk_label_zh,
        "risk_label_en": risk_label_en,
        "recommendation_zh": recommendation_zh,
        "recommendation_en": recommendation_en,
        "input_summary_rows": input_summary_rows,
    }


# ─────────────────────────────────────────────────────────────
# 登入 / 登出
# ─────────────────────────────────────────────────────────────
def login_view(request):
    """共用密碼登入頁面。"""
    if request.session.get("authenticated"):
        return redirect("prediction:prediction_page")

    error = False
    if request.method == "POST":
        password = request.POST.get("password", "")
        demo_pw = getattr(settings, "DPM_ACCESS_PASSWORD", "dpm2026")
        full_pw = getattr(settings, "DPM_FULL_PASSWORD", "")
        if full_pw and password == full_pw:
            request.session["authenticated"] = True
            request.session["access_level"] = "full"
            request.session["lang"] = request.POST.get("lang", "zh")
            return redirect("prediction:prediction_page")
        elif password == demo_pw:
            request.session["authenticated"] = True
            request.session["access_level"] = "demo"
            request.session["lang"] = request.POST.get("lang", "zh")
            return redirect("prediction:prediction_page")
        error = True

    return render(request, "prediction/login.html", {"error": error})


def logout_view(request):
    """登出並清除 session。"""
    request.session.flush()
    return redirect("prediction:login")


# ─────────────────────────────────────────────────────────────
# Views
# ─────────────────────────────────────────────────────────────
def prediction_page(request):
    """主預測頁面：GET 顯示表單，POST 處理預測。"""
    result = None
    error_msg = None

    lang = request.session.get("lang", "zh")

    if request.method == "POST":
        lang = request.POST.get("lang", lang)
        request.session["lang"] = lang
        form = PredictionForm(request.POST)
        if form.is_valid():
            try:
                result = predict_model(form.cleaned_data, lang=lang)
                # 存入 session 以供下載使用
                request.session["last_prediction"] = {
                    "result": result,
                    "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                }
            except Exception as e:
                error_msg = f"預測過程發生錯誤：{e}"
    else:
        form = PredictionForm()

    return render(request, "prediction/prediction_page.html", {
        "form": form,
        "result": result,
        "result_json": json.dumps(result, ensure_ascii=False) if result else "null",
        "error_msg": error_msg,
        "session_lang": request.session.get("lang", "zh"),
    })


# ─────────────────────────────────────────────────────────────
# 欄位分級：依 SHAP 重要性 + 模型程式碼是否有缺失保護
# ─────────────────────────────────────────────────────────────
# 【必填】模型程式碼直接存取、無 guard，缺少會 crash
CRITICAL_COLUMNS = [
    "loan term",                          # SHAP 0.234 — 用於 payment_progress_ratio
    "paid installments",                  # SHAP 0.147 — 用於 payment_progress_ratio
    "post code of residential address",   # SHAP 0.212 — WoE 編碼
    "education",                          # SHAP 0.065 — WoE 編碼
    "main business",                      # SHAP 0.057 — WoE 編碼
    "job tenure",                         # SHAP 0.049 — 用於 job_stable
    "month salary",                       # SHAP 0.032 — 用於 payment_pressure
    "post code of permanent address",     # SHAP 0.008 — 用於 address_match
    "residence status",                   # SHAP 0.005 — WoE 編碼
    "product",                            # SHAP ≈0    — WoE 編碼（模型必要）
]

# 【選填】模型程式碼有 `if col in df.columns` 保護，缺少不會 crash
# 但會影響預測品質，缺少時預設填 0 並提示使用者
OPTIONAL_COLUMNS = {
    # 欄位名: (預設值, SHAP 重要性等級)
    "number of overdue before the first month":               (0, "★★★ 極高"),
    "number of overdue in the first half of the first month": (0, "★★★ 極高"),
    "number of overdue in the second half of the first month":(0, "★★★ 極高"),
    "number of overdue in the second month":                  (0, "★★★ 極高"),
    "number of overdue in the third month":                   (0, "★★★ 極高"),
    "number of overdue in the fourth month":                  (0, "★★★ 極高"),
    "number of overdue in the fifth month":                   (0, "★★★ 極高"),
    "number of overdue in the sixth month":                   (0, "★★★ 極高"),
    "debt_to_income_ratio":                                   (0, "★☆☆ 低"),
    "payment_to_income_ratio":                                (0, "★☆☆ 低"),
}


def upload_predict(request):
    """上傳 Excel/CSV 批次預測。"""
    if request.method != "POST" or "file" not in request.FILES:
        return HttpResponse("請上傳檔案。", status=400)

    uploaded = request.FILES["file"]
    ext = os.path.splitext(uploaded.name)[1].lower()

    try:
        if ext == ".csv":
            df = pd.read_csv(uploaded)
        elif ext in (".xlsx", ".xls"):
            # 嘗試找到含有正確欄位的工作表
            xls = pd.ExcelFile(uploaded, engine="openpyxl")
            df = None
            # 優先讀取「空白範本」或第一個含有必填欄位的 sheet
            preferred_sheets = ["空白範本", "填寫範例"]
            for sheet in preferred_sheets + xls.sheet_names:
                if sheet not in xls.sheet_names:
                    continue
                candidate = pd.read_excel(xls, sheet_name=sheet)
                candidate.columns = [str(c).strip() for c in candidate.columns]
                if any(col in candidate.columns for col in CRITICAL_COLUMNS):
                    df = candidate
                    break
            if df is None:
                # 都找不到就讀第一個 sheet
                df = pd.read_excel(xls, sheet_name=0)
        else:
            return HttpResponse("僅支援 .xlsx / .xls / .csv 格式。", status=400)
    except Exception as e:
        return HttpResponse(f"檔案讀取失敗：{e}", status=400)

    # 欄位驗證：清理欄位名稱（去除前後空白和換行）
    df.columns = [col.strip() for col in df.columns]

    # 1) 必填欄位：缺少 → 擋下
    missing_critical = [col for col in CRITICAL_COLUMNS if col not in df.columns]
    if missing_critical:
        lang = request.session.get("lang", "zh")
        missing_str = "、".join(missing_critical)
        bullet_missing = "\n".join(f"  • {c}" for c in missing_critical)
        found_cols = ", ".join(df.columns.tolist())
        if lang == "en":
            error_msg = (
                f"⚠ Upload failed — {len(missing_critical)} required column(s) not found:\n"
                f"{bullet_missing}\n\n"
                f"Columns detected in your file:\n  {found_cols}\n\n"
                f"Please download the template and re-upload with all required columns."
            )
        else:
            error_msg = (
                f"⚠ 上傳失敗 — 缺少 {len(missing_critical)} 個必填欄位：\n"
                f"{bullet_missing}\n\n"
                f"您的檔案現有欄位：\n  {found_cols}\n\n"
                f"請下載範本檔案，補齊所有必填欄位後重新上傳。"
            )
        return render(request, "prediction/prediction_page.html", {
            "form": PredictionForm(),
            "result_json": "null",
            "error_msg": error_msg,
            "session_lang": lang,
        })

    # 2) 選填欄位：缺少 → 填預設值 + 記錄警告
    warning_lines = []
    for col, (default_val, importance) in OPTIONAL_COLUMNS.items():
        if col not in df.columns:
            df[col] = default_val
            warning_lines.append(f"• {col}（重要性：{importance}，已預設為 {default_val}）")

    warning_msg = None
    if warning_lines:
        warning_msg = (
            f"⚠ 以下 {len(warning_lines)} 個選填欄位未提供，已自動填入預設值。\n"
            "若為新客戶（無歷史逾期記錄），預設值 0 為合理值；\n"
            "若為回頭客，缺少逾期欄位將顯著影響預測準確度。\n\n"
            + "\n".join(warning_lines)
        )

    # ── Demo 模式筆數限制 ──
    DEMO_ROW_LIMIT = 10
    access_level = request.session.get("access_level", "demo")
    if access_level != "full" and len(df) > DEMO_ROW_LIMIT:
        lang = request.session.get("lang", "zh")
        if lang == "en":
            limit_msg = (
                f"Demo mode supports up to {DEMO_ROW_LIMIT} records per upload. "
                f"Your file contains {len(df)} records. "
                f"Please contact us for full-access credentials if you need batch predictions at scale."
            )
        else:
            limit_msg = (
                f"Demo 模式每次最多上傳 {DEMO_ROW_LIMIT} 筆，您的檔案共有 {len(df)} 筆。\n"
                f"如需批次大量預測，請聯絡管理員取得完整授權帳號。"
            )
        return render(request, "prediction/prediction_page.html", {
            "form": PredictionForm(),
            "result_json": "null",
            "error_msg": limit_msg,
            "session_lang": lang,
        })

    # 從 POST 取得當前語言（前端 toggle 可能與登入時不同），並更新 session
    lang = request.POST.get("lang", request.session.get("lang", "zh"))
    request.session["lang"] = lang

    # Translate English categorical values → Chinese for model processing
    for col in EN_CATEGORICAL_COLS:
        if col in df.columns:
            df[col] = df[col].map(lambda v: EN_TO_ZH.get(str(v).strip(), v) if pd.notna(v) else v)

    # Validate categorical values — catch unmapped values before model crashes
    VALID_ZH_VALUES = {
        "education":        set(v for _, v in EN_TO_ZH.items() if _ in ["Grad+","Graduate","Master","College/Univ","University","College","HS/Voc","High School","Elementary","Other","Others"]),
        "residence status": {"自有", "租屋", "配偶", "親屬", "宿舍"},
        "main business":    {v for k, v in EN_TO_ZH.items() if k in ["Manufacturing","Services","Service","Commerce","Retail","Tech","Finance","Insurance","Securities","Gov/Education","Military","F&B","Transport","Construction","Real Estate","Warehousing","Telecom","Utilities","Agriculture","Fishery","Mining","Professional","Freelance","E-commerce","Student","Homemaker","Social/Personal","Other","Others"]},
        "product":          {"瘦身美容", "3C家電", "個人用品", "其他"},
    }
    invalid_rows = {}
    for col, valid_set in VALID_ZH_VALUES.items():
        if col in df.columns:
            bad_mask = df[col].apply(lambda v: pd.notna(v) and str(v).strip() not in valid_set)
            bad_vals = df.loc[bad_mask, col].unique().tolist()
            if bad_vals:
                invalid_rows[col] = bad_vals
    if invalid_rows:
        bullet = "\n".join(f"  • {col}：{vals}" for col, vals in invalid_rows.items())
        if lang == "en":
            err = (
                f"⚠ Upload failed — unrecognised categorical values found:\n{bullet}\n\n"
                f"Please check the template for valid values."
            )
        else:
            err = (
                f"⚠ 上傳失敗 — 以下欄位包含不合法的分類值：\n{bullet}\n\n"
                f"請參考範本中的合法選項後重新上傳。"
            )
        return render(request, "prediction/prediction_page.html", {
            "form": PredictionForm(),
            "result_json": "null",
            "error_msg": err,
            "session_lang": lang,
        })

    # Convert USD → NTD for English users (model expects NTD)
    if lang == "en" and "month salary" in df.columns:
        df["month salary"] = (df["month salary"] / NTD_TO_USD).round(0).astype(int)

    predictor = _get_predictor()

    # 保留原始資料中的非模型欄位（如 name, ID 等）
    all_model_cols = set(CRITICAL_COLUMNS) | set(OPTIONAL_COLUMNS.keys())
    extra_cols = [c for c in df.columns if c not in all_model_cols]
    extra_df = df[extra_cols].reset_index(drop=True) if extra_cols else None

    # 切換工作目錄以便相對路徑正確
    project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
    prediction_dir = os.path.join(project_root, "Prediction")
    original_cwd = os.getcwd()
    os.chdir(prediction_dir)
    try:
        result_df = predictor.predict_with_details(df, simplified_output=True)
    except Exception as e:
        return render(request, "prediction/prediction_page.html", {
            "form": PredictionForm(),
            "result_json": "null",
            "error_msg": f"批次預測失敗：{e}",
            "session_lang": request.session.get("lang", "zh"),
        })
    finally:
        os.chdir(original_cwd)

    # 將非模型欄位（name 等）合併到結果最前面
    result_df = result_df.reset_index(drop=True)
    if extra_df is not None:
        result_df = pd.concat([extra_df, result_df], axis=1)

    # 移除內部技術欄位（業務使用者不需要看）
    drop_cols = ["predicted_default", "predicted_default_optimal", "threshold_difference"]
    result_df = result_df.drop(columns=[c for c in drop_cols if c in result_df.columns])

    # Translate model output values to English if session lang is "en"
    if lang == "en":
        if "risk_grade" in result_df.columns:
            result_df["risk_grade"] = result_df["risk_grade"].map(
                lambda v: ZH_TO_EN_GRADE.get(str(v), v))
        if "risk_action" in result_df.columns:
            result_df["risk_action"] = result_df["risk_action"].map(
                lambda v: ZH_TO_EN_ACTION.get(str(v), v))
        if "risk_action_optimal" in result_df.columns:
            result_df["risk_action_optimal"] = result_df["risk_action_optimal"].map(
                lambda v: ZH_TO_EN_ACTION.get(str(v), v))
        # Also translate categorical input columns back to English for the output
        for col in EN_CATEGORICAL_COLS:
            if col in result_df.columns:
                result_df[col] = result_df[col].map(
                    lambda v: ZH_TO_EN.get(str(v), v) if pd.notna(v) else v)

    # 存入 session 供下載
    batch_records = result_df.to_dict(orient="records")
    batch_columns = list(result_df.columns)
    request.session["batch_result"] = {
        "records": json.loads(json.dumps(batch_records, default=str)),
        "columns": batch_columns,
        "count": len(result_df),
        "filename": uploaded.name,
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }

    # 統計摘要
    summary = {
        "total": len(result_df),
        "filename": uploaded.name,
    }
    if "risk_grade" in result_df.columns:
        summary["grade_counts"] = result_df["risk_grade"].value_counts().to_dict()
    if "default_probability" in result_df.columns:
        summary["avg_prob"] = round(float(result_df["default_probability"].mean()), 2)
        summary["max_prob"] = round(float(result_df["default_probability"].max()), 2)
        summary["min_prob"] = round(float(result_df["default_probability"].min()), 2)

    return render(request, "prediction/prediction_page.html", {
        "form": PredictionForm(),
        "batch_summary": summary,
        "result_json": "null",
        "warning_msg": warning_msg,
        "session_lang": request.session.get("lang", "zh"),
    })


def download_batch_result(request):
    """下載批次預測結果 Excel（含格式美化）。"""
    batch = request.session.get("batch_result")
    if not batch:
        return HttpResponse("尚無批次預測結果可下載。", status=400)

    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    result_df = pd.DataFrame(batch["records"], columns=batch["columns"])
    timestamp = batch["timestamp"].replace(":", "").replace(" ", "_")
    lang = request.session.get("lang", "zh")

    # 欄位名稱對照
    if lang == "en":
        COL_LABELS = {
            "name": "Name",
            "education": "Education",
            "month salary": "Monthly Salary",
            "job tenure": "Job Tenure (yrs)",
            "residence status": "Residence Status",
            "main business": "Industry",
            "product": "Loan Purpose",
            "loan term": "Loan Term (months)",
            "paid installments": "Paid Installments",
            "post code of permanent address": "Perm. Address Postal Code",
            "post code of residential address": "Res. Address Postal Code",
            "debt_to_income_ratio": "Debt-to-Income Ratio",
            "payment_to_income_ratio": "Payment-to-Income Ratio",
            "default_probability": "Default Probability (%)",
            "risk_score": "Risk Score",
            "risk_grade": "Risk Grade",
            "risk_action": "Recommendation",
            "risk_action_optimal": "Recommendation (Optimal)",
            "risk_alert": "Risk Alert",
        }
    else:
        COL_LABELS = {
            "name": "姓名",
            "default_probability": "違約機率 (%)",
            "risk_score": "風險分數",
            "risk_grade": "風險等級",
            "risk_action": "建議行動",
            "risk_action_optimal": "建議行動（最佳）",
            "risk_alert": "風險警示",
        }

    sheet_name = "Prediction Results" if lang == "en" else "預測結果"
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        result_df.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]

        # ── 樣式定義 ──
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="C2410C", end_color="C2410C", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_align_center = Alignment(horizontal="center", vertical="center")
        cell_align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_border = Border(
            bottom=Side(style="thin", color="D1D5DB"),
        )

        # 風險等級底色
        grade_fills = {
            "A": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),  # emerald
            "B": PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),  # green
            "C": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),  # amber
            "D": PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid"),  # orange
            "E": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),  # red
        }

        # 需要靠左對齊的欄位（文字較長）
        left_align_cols = {"risk_action", "risk_action_optimal", "risk_alert", "name"}

        num_cols = len(result_df.columns)
        num_rows = len(result_df)

        # ── 表頭格式 + 中文名 ──
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col_idx)
            orig_name = cell.value
            cell.value = COL_LABELS.get(orig_name, orig_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # ── 資料列格式 ──
        for row_idx in range(2, num_rows + 2):
            for col_idx in range(1, num_cols + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                col_name = result_df.columns[col_idx - 1]

                # 對齊
                if col_name in left_align_cols:
                    cell.alignment = cell_align_left
                else:
                    cell.alignment = cell_align_center

                cell.border = thin_border

            # 整列依風險等級上底色
            grade_col_idx = list(result_df.columns).index("risk_grade") + 1 if "risk_grade" in result_df.columns else None
            if grade_col_idx:
                grade_val = str(ws.cell(row=row_idx, column=grade_col_idx).value or "")
                grade_letter = grade_val[0] if grade_val else ""
                row_fill = grade_fills.get(grade_letter)
                if row_fill:
                    for col_idx in range(1, num_cols + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = row_fill

        # ── 欄寬自動調整 ──
        for col_idx in range(1, num_cols + 1):
            col_name = result_df.columns[col_idx - 1]
            header_len = len(str(COL_LABELS.get(col_name, col_name)))
            max_len = max(header_len, 8)  # 最少 8
            # 抽樣前 20 行估算寬度
            for row_idx in range(2, min(num_rows + 2, 22)):
                val = str(ws.cell(row=row_idx, column=col_idx).value or "")
                max_len = max(max_len, len(val))
            # 中文字寬度 ×1.8，上限 40
            width = min(max_len * 1.8 + 2, 40)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # ── 凍結首列 ──
        ws.freeze_panes = "A2"

    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="batch_prediction_{timestamp}.xlsx"'
    )
    return response


def download_csv(request):
    """下載 CSV 格式預測報表。"""
    prediction = request.session.get("last_prediction")
    if not prediction:
        return HttpResponse("No prediction result available.", status=400)

    result = prediction["result"]
    timestamp = prediction["timestamp"]
    lang = request.session.get("lang", "zh")

    response = HttpResponse(content_type="text/csv; charset=utf-8-sig")
    response["Content-Disposition"] = (
        f'attachment; filename="prediction_report_{timestamp.replace(":", "").replace(" ", "_")}.csv"'
    )

    writer = csv.writer(response)

    if lang == "en":
        writer.writerow(["Default Prediction Report"])
        writer.writerow(["Generated at", timestamp])
        writer.writerow([])
        writer.writerow(["== Prediction Results =="])
        writer.writerow(["Default Probability", f'{result["default_probability"]}%'])
        writer.writerow(["Risk Score", result["risk_score"]])
        writer.writerow(["Risk Grade", result["risk_grade"]])
        writer.writerow(["Risk Label", result["risk_label"]])
        writer.writerow(["Risk Alert", result["risk_alert"]])
        writer.writerow(["Recommendation", result["recommendation"]])
        writer.writerow([])
        writer.writerow(["== Input Data Summary =="])
    else:
        writer.writerow(["違約預測報表"])
        writer.writerow(["產出時間", timestamp])
        writer.writerow([])
        writer.writerow(["== 預測結果 =="])
        writer.writerow(["違約機率", f'{result["default_probability"]}%'])
        writer.writerow(["風險分數", result["risk_score"]])
        writer.writerow(["風險等級", result["risk_grade"]])
        writer.writerow(["風險標籤", result["risk_label"]])
        writer.writerow(["風險警示", result["risk_alert"]])
        writer.writerow(["建議行動", result["recommendation"]])
        writer.writerow([])
        writer.writerow(["== 輸入資料摘要 =="])

    for key, value in result["input_summary"].items():
        writer.writerow([key, value])

    return response


def download_excel(request):
    """下載 Excel 格式預測報表。"""
    prediction = request.session.get("last_prediction")
    if not prediction:
        return HttpResponse("No prediction result available.", status=400)

    result = prediction["result"]
    timestamp = prediction["timestamp"]
    lang = request.session.get("lang", "zh")

    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return HttpResponse("Missing openpyxl package.", status=500)

    wb = openpyxl.Workbook()
    ws = wb.active

    # 標題
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="E65100", end_color="E65100", fill_type="solid")
    ws.merge_cells("A1:B1")
    ws["A1"].font = header_font
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center")

    section_font = Font(bold=True, size=11)

    if lang == "en":
        ws.title = "Prediction Report"
        ws["A1"] = "Default Prediction Report"
        ws["A2"] = "Generated at"
        ws["B2"] = timestamp
        ws["A4"] = "Prediction Results"
        ws["A4"].font = section_font
        rows = [
            ("Default Probability", f'{result["default_probability"]}%'),
            ("Risk Score", result["risk_score"]),
            ("Risk Grade", result["risk_grade"]),
            ("Risk Label", result["risk_label"]),
            ("Risk Alert", result["risk_alert"]),
            ("Recommendation", result["recommendation"]),
        ]
    else:
        ws.title = "違約預測報表"
        ws["A1"] = "違約預測報表"
        ws["A2"] = "產出時間"
        ws["B2"] = timestamp
        ws["A4"] = "預測結果"
        ws["A4"].font = section_font
        rows = [
            ("違約機率", f'{result["default_probability"]}%'),
            ("風險分數", result["risk_score"]),
            ("風險等級", result["risk_grade"]),
            ("風險標籤", result["risk_label"]),
            ("風險警示", result["risk_alert"]),
            ("建議行動", result["recommendation"]),
        ]

    for i, (k, v) in enumerate(rows, start=5):
        ws[f"A{i}"] = k
        ws[f"B{i}"] = v

    # 輸入摘要
    row_offset = 5 + len(rows) + 1
    ws[f"A{row_offset}"] = "Input Data Summary" if lang == "en" else "輸入資料摘要"
    ws[f"A{row_offset}"].font = section_font
    for i, (k, v) in enumerate(result["input_summary"].items(), start=row_offset + 1):
        ws[f"A{i}"] = k
        ws[f"B{i}"] = v

    # 欄寬
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 40

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="prediction_report_{timestamp.replace(":", "").replace(" ", "_")}.xlsx"'
    )
    return response


def download_template(request):
    """下載批次預測範本 Excel（中英雙語版）。"""
    from io import BytesIO

    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    lang = request.GET.get("lang", request.session.get("lang", "zh"))
    is_zh = lang == "zh"

    # ── 欄位順序（以使用者直覺順序排列，非 SHAP 順序）──
    TEMPLATE_COLS = [
        "education",
        "month salary",
        "job tenure",
        "residence status",
        "main business",
        "product",
        "loan term",
        "paid installments",
        "post code of permanent address",
        "post code of residential address",
        "debt_to_income_ratio",
        "payment_to_income_ratio",
        "number of overdue before the first month",
        "number of overdue in the first half of the first month",
        "number of overdue in the second half of the first month",
        "number of overdue in the second month",
        "number of overdue in the third month",
        "number of overdue in the fourth month",
        "number of overdue in the fifth month",
        "number of overdue in the sixth month",
    ]

    REQUIRED_SET = set(CRITICAL_COLUMNS)

    # 欄位中文說明
    ZH_DESC = {
        "education":                                              ("教育程度", "Graduate / University / High School"),
        "month salary":                                           ("月薪（元）", "數字，例如：50000"),
        "job tenure":                                             ("工作年資（年）", "數字，例如：3.5"),
        "residence status":                                       ("居住狀態", "Owned / Rented / Family / Mortgage"),
        "main business":                                          ("行業別", "Finance / Retail / Tech / Manufacturing / Service / Others"),
        "product":                                                ("借款目的", "Personal Loan / Credit Card / Car Loan"),
        "loan term":                                              ("貸款期數（月）", "整數，例如：24"),
        "paid installments":                                      ("已繳期數", "整數，例如：6"),
        "post code of permanent address":                         ("戶籍郵遞區號", "數字，例如：100"),
        "post code of residential address":                       ("居住郵遞區號", "數字，例如：235"),
        "debt_to_income_ratio":                                   ("負債收入比", "0–1 之間，例如：0.35"),
        "payment_to_income_ratio":                                ("還款收入比", "0–1 之間，例如：0.20"),
        "number of overdue before the first month":               ("第1月前逾期次數", "整數，新客戶填 0"),
        "number of overdue in the first half of the first month": ("第1月上半逾期次數", "整數，新客戶填 0"),
        "number of overdue in the second half of the first month":("第1月下半逾期次數", "整數，新客戶填 0"),
        "number of overdue in the second month":                  ("第2月逾期次數", "整數，新客戶填 0"),
        "number of overdue in the third month":                   ("第3月逾期次數", "整數，新客戶填 0"),
        "number of overdue in the fourth month":                  ("第4月逾期次數", "整數，新客戶填 0"),
        "number of overdue in the fifth month":                   ("第5月逾期次數", "整數，新客戶填 0"),
        "number of overdue in the sixth month":                   ("第6月逾期次數", "整數，新客戶填 0"),
    }

    EN_DESC = {
        "education":                                              ("Education Level", "Graduate / Master / University / High School / Other"),
        "month salary":                                           ("Monthly Salary", "Number, e.g. 50000"),
        "job tenure":                                             ("Job Tenure (yrs)", "Number, e.g. 3.5"),
        "residence status":                                       ("Residence Status", "Owned / Rented / Family / Spouse / Dormitory"),
        "main business":                                          ("Industry", "Finance / Tech / Manufacturing / Services / Commerce / Others"),
        "product":                                                ("Loan Purpose", "Personal / Beauty / 3C/Appliance / Others"),
        "loan term":                                              ("Loan Term (months)", "Integer, e.g. 24"),
        "paid installments":                                      ("Paid Installments", "Integer, e.g. 6"),
        "post code of permanent address":                         ("Perm. Address Postal Code", "Number, e.g. 100"),
        "post code of residential address":                       ("Res. Address Postal Code", "Number, e.g. 235"),
        "debt_to_income_ratio":                                   ("Debt-to-Income Ratio", "0–1, e.g. 0.35"),
        "payment_to_income_ratio":                                ("Payment-to-Income Ratio", "0–1, e.g. 0.20"),
        "number of overdue before the first month":               ("Overdue Count Before Month 1", "Integer, 0 for new clients"),
        "number of overdue in the first half of the first month": ("Overdue Count Month 1 First Half", "Integer, 0 for new clients"),
        "number of overdue in the second half of the first month":("Overdue Count Month 1 Second Half", "Integer, 0 for new clients"),
        "number of overdue in the second month":                  ("Overdue Count Month 2", "Integer, 0 for new clients"),
        "number of overdue in the third month":                   ("Overdue Count Month 3", "Integer, 0 for new clients"),
        "number of overdue in the fourth month":                  ("Overdue Count Month 4", "Integer, 0 for new clients"),
        "number of overdue in the fifth month":                   ("Overdue Count Month 5", "Integer, 0 for new clients"),
        "number of overdue in the sixth month":                   ("Overdue Count Month 6", "Integer, 0 for new clients"),
    }

    SAMPLE_ROWS = [
        ["Graduate",     150000, 12,  "Owned",    "Finance",       "Personal",     12, 6,  100, 100, 0.15, 0.10, 0, 0, 0, 0, 0, 0, 0, 0],
        ["University",    32000,  2,  "Rented",   "Commerce",      "Beauty",        36, 12, 235, 235, 0.45, 0.30, 1, 0, 0, 0, 0, 0, 0, 0],
        ["High School",   28000,  0.5,"Family",   "Others",        "Personal",      60, 3,  400, 405, 0.65, 0.40, 2, 1, 1, 1, 0, 0, 0, 0],
        ["University",    85000,  8,  "Owned",    "Tech",          "3C/Appliance",  48, 24, 110, 110, 0.25, 0.15, 0, 0, 0, 0, 0, 0, 0, 0],
        ["Graduate",     120000,  5,  "Owned",    "Tech",          "Personal",      24, 10, 300, 300, 0.35, 0.20, 1, 0, 0, 0, 0, 0, 0, 0],
        ["High School",   42000,  3,  "Rented",   "Services",      "Beauty",        12, 2,  800, 802, 0.55, 0.35, 0, 0, 0, 0, 0, 0, 0, 0],
        ["University",    30000,  1,  "Family",   "Commerce",      "Personal",      36, 1,  900, 900, 0.75, 0.45, 3, 2, 2, 2, 1, 0, 0, 0],
        ["University",    65000, 12,  "Owned",    "Manufacturing", "3C/Appliance",  72, 48, 500, 500, 0.20, 0.12, 0, 0, 0, 0, 0, 0, 0, 0],
        ["High School",   25000,  4,  "Rented",   "Services",      "Personal",      24, 18, 700, 700, 0.40, 0.25, 0, 1, 0, 0, 0, 0, 0, 0],
        ["Graduate",      95000,  6,  "Spouse",   "Finance",       "Beauty",        48, 36, 200, 200, 0.10, 0.08, 0, 0, 0, 0, 0, 0, 0, 0],
    ]

    wb = openpyxl.Workbook()

    # ────────────────────────────────────────────────
    # Sheet 1: 填寫範例 / Sample Data
    # ────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "填寫範例" if is_zh else "Sample Data"

    navy_fill   = PatternFill("solid", fgColor="1A3A6C")
    orange_fill = PatternFill("solid", fgColor="E65100")
    yellow_fill = PatternFill("solid", fgColor="FFF3E0")
    white_font  = Font(bold=True, color="FFFFFF", size=10)
    thin_side   = Side(style="thin", color="DDDDDD")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    desc_map = ZH_DESC if is_zh else EN_DESC

    # Row 1: English column names (system-required)
    for c_idx, col in enumerate(TEMPLATE_COLS, 1):
        cell = ws1.cell(row=1, column=c_idx, value=col)
        cell.fill = navy_fill if col in REQUIRED_SET else orange_fill
        cell.font = white_font
        cell.alignment = center_align
        cell.border = thin_border

    # Row 2 (ZH only): Chinese label hints
    data_start_row = 2
    if is_zh:
        zh_hint_fill = PatternFill("solid", fgColor="EBF3FB")
        zh_hint_font = Font(size=9, color="1A3A6C", italic=True)
        for c_idx, col in enumerate(TEMPLATE_COLS, 1):
            zh_name, _ = ZH_DESC[col]
            cell = ws1.cell(row=2, column=c_idx, value=zh_name)
            cell.fill = zh_hint_fill
            cell.font = zh_hint_font
            cell.alignment = center_align
            cell.border = thin_border
        ws1.row_dimensions[2].height = 20
        data_start_row = 3

    for r_idx, row_data in enumerate(SAMPLE_ROWS, data_start_row):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            if r_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F9FAFB")

    for c_idx, col in enumerate(TEMPLATE_COLS, 1):
        max_len = max(
            len(str(ws1.cell(row=r, column=c_idx).value or ""))
            for r in range(1, data_start_row + len(SAMPLE_ROWS))
        )
        ws1.column_dimensions[get_column_letter(c_idx)].width = min(max_len + 3, 32)
    ws1.row_dimensions[1].height = 36

    # Legend note
    note_row = len(SAMPLE_ROWS) + 3
    ws1.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=len(TEMPLATE_COLS))
    note_cell = ws1.cell(row=note_row, column=1)
    if is_zh:
        note_cell.value = "■ 深藍欄位 = 必填（缺少將無法預測）　■ 橘色欄位 = 選填（缺少自動補 0，但影響準確度）　■ 逾期欄位對模型貢獻 >80%，回頭客請務必填寫"
    else:
        note_cell.value = "■ Navy = Required (prediction will fail without these)   ■ Orange = Optional (defaults to 0 if missing, affects accuracy)   ■ Overdue fields contribute >80% to model accuracy — always fill for returning clients"
    note_cell.font = Font(italic=True, size=9, color="555555")
    note_cell.alignment = Alignment(wrap_text=True)

    # ────────────────────────────────────────────────
    # Sheet 2: 欄位說明 / Field Guide
    # ────────────────────────────────────────────────
    ws2 = wb.create_sheet("欄位說明" if is_zh else "Field Guide")

    guide_headers = (
        ["欄位名稱（系統用）", "中文名稱", "說明 / 範例值", "必填？"] if is_zh
        else ["Column Name (system)", "Display Name", "Description / Sample Values", "Required?"]
    )
    for c_idx, h in enumerate(guide_headers, 1):
        cell = ws2.cell(row=1, column=c_idx, value=h)
        cell.fill = navy_fill
        cell.font = white_font
        cell.alignment = center_align
        cell.border = thin_border

    for r_idx, col in enumerate(TEMPLATE_COLS, 2):
        disp_name, desc_val = desc_map[col]
        req_label = ("✔ 必填" if is_zh else "✔ Required") if col in REQUIRED_SET else ("選填" if is_zh else "Optional")
        row_vals = [col, disp_name, desc_val, req_label]
        req_color = "FFF3E0" if col not in REQUIRED_SET else "E8F0FE"
        for c_idx, val in enumerate(row_vals, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            cell.fill = PatternFill("solid", fgColor=req_color)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws2.column_dimensions["A"].width = 52
    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["C"].width = 40
    ws2.column_dimensions["D"].width = 12
    ws2.row_dimensions[1].height = 28

    # ── Return response ──
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    from urllib.parse import quote
    filename = "預測範本.xlsx" if is_zh else "Prediction_Template.xlsx"
    encoded  = quote(filename)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f"attachment; filename=\"{filename}\"; filename*=UTF-8''{encoded}"
    return response
